#coding=utf-8
from openpyxl import load_workbook
class doExcel:
    def __init__(self,file_name,sheet_name):
        self.file_name=file_name
        self.sheet_name=sheet_name
    def read_excel(self):
        wb=load_workbook(self.file_name)
        sheet=wb[self.sheet_name]
        test_data=[]
        for row in range(2,sheet.max_row+1):
            row_data={}
            row_data["caseid"]=sheet.cell(row,1).value
            row_data["module"] = sheet.cell(row,2).value
            row_data["description"] = sheet.cell(row,3).value
            row_data["method"] = sheet.cell(row,4).value
            row_data["url"] = sheet.cell(row,5).value
            if "{mobile}" in sheet.cell(row,6).value:
                row_data["params"]=sheet.cell(row,6).value.replace("{mobile}",str(self.read_moblie()))
                self.update_mobile(self.read_moblie()+1)
            else:
                row_data["params"] = sheet.cell(row,6).value
            row_data["expextedresult"]=sheet.cell(row,7).value
            test_data.append(row_data)
        wb.close()
        return test_data



    def read_moblie(self):
        wb=load_workbook(self.file_name)
        sheet=wb["mobile"]
        tel=sheet.cell(1,2).value
        wb.close()
        return tel

    def update_mobile(self,new_tel):
        wb=load_workbook(self.file_name)
        sheet=wb["mobile"]
        sheet.cell(1,2).value=new_tel
        wb.save(self.file_name)
        wb.close()




if __name__ == '__main__':
    de=doExcel("../test_case/api_case.xlsx","case")
    # print(de.update_mobile(),type(de.update_mobile()))
    print(de.read_excel())




